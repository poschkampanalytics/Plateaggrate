# ==============================================
# Author: Broder Poschkamp
# E-Mail: broder.poschkamp@gmail.com
# Program: Plateaggrate
# Date: 190926
# Comments: program is controlled with GUI, shell is for better understanding
#     every widget has a function, which usually controls input/ output
#     everything is managed by the function called: "action()"
#     takes other functions to handle every single step, which is explained there
# (c) MIT License
# ==============================================
# imports
import datetime
import imghdr
import os
import shutil
import sys
from tkinter import Button, Label, Text, Scrollbar, Entry, messagebox, filedialog, Tk, StringVar, BooleanVar
from tkinter import INSERT, Checkbutton, PhotoImage, CENTER, VERTICAL, HORIZONTAL, ttk, END

try:
    import numpy as np
    import cv2
    import tensorflow as tf
    from openpyxl import *
except ImportError as err:
    sys.exc_info()
    print("""It's not possible to load all used modules. Please read README.txt for more information.""")


# ==============================================
# general
__author__ = "Broder Poschkamp"
__credits__ = "Broder Poschkamp"
__license__ = "MIT Licence"
__version__ = "1.0.1"
__email__ = "broder.poschkamp@gmail.com"
__status__ = "development"

# ==============================================
# functions
#    all functions are used together in "action"; it's the "run analysis" -button command

# =======================
# general functions
#    directory for saving files; delete if it is already there


def make_dir(input_path, name="data", nr=0):
    """
    This function creates a folder, which is usually named "data".
    It has a recursive form, to name folders differently: it appends "_[Nr]"; Nr starts with 1,
    parameter:
    path:           Path where a new folder is made.
    nr:             Number of folder, which is named the same
    return:
    dir path:       returns the path for saving files
    """
    # make sure that folder isn't named the same
    #   first run
    if nr == 0:
        dir_path = os.path.join(input_path, name)
        if os.path.isdir(dir_path):
            nr = 1
            return make_dir(dir_path, nr=nr)
        else:
            # general chase
            os.mkdir(dir_path)
            return dir_path

    elif nr == 1:
        dir_path = input_path + "_" + str(nr)
        if os.path.isdir(dir_path):
            nr += 1
            return make_dir(dir_path, nr=nr)
        else:
            os.mkdir(dir_path)
            return dir_path

    else:
        len_nr = len(str(nr))
        dir_path = input_path[:-len_nr] + str(nr)
        if os.path.isdir(dir_path):
            nr += 1
            return make_dir(dir_path, nr=nr)
        else:
            os.mkdir(dir_path)
            return dir_path


# get images and data type
def get_images_list(path, show_dtype=False):
    """ The function returns all images names in a directory.
    It also checks the dtype of the first image (optional).
    parameter:
    path:               Path of directory, where images are saved
    show_dtype:         Change to True, to return also dtype as string
    return:
    l_output_files:     list of all images files (as string)
    example_dtype:      the data type of the first image as string
    """
    l_files = os.listdir(path)
    l_output_files = []
    if len(l_files):  # check if list isn't empty
        for filename in l_files:
            if filename.lower().endswith((".png", ".jpg", ".jpeg", ".tiff", ".tif")):
                l_output_files.append(filename)
    else:
        # return error; if there aren't images in folder
        error_input_images()

    # check data type
    if show_dtype:
        example_file = os.path.join(path, l_output_files[0])
        example_dtype = imghdr.what(example_file)
        return l_output_files, example_dtype
    else:
        return l_output_files


# =======================
# openpyxl/ excel

def len_row(sheet):
    """ This function is designed for getting the number of rows in an excel sheet, which is opened with openpyxl.
    parameter:
    sheet:              The active sheet which is open in openpyxl.
    return:
    row_number:         The Number of rows in an excel sheet."""
    row_number = len(list(sheet.rows))
    return row_number

# =======================
# Image pipeline

# Image Processing
#    complete Images processing is done with function: image_processing
#    small functions for single steps; better for debugging; individualisation


def shape(img, dim=(60, 60)):
    """
    Moving a smaller image to the middle of an Area with a given size (dim).
    The background colour is the mean of the first row.
    parameter:
    img:                image which was loaded with cv2; image has one dimension (e.g. grayscale)
    dim:                size of output image, (standard size is 60x 60 pixel)
    Background: change background to mean of first row
    Edge: color edge to mean
    """
    # Size
    xsize, ysize = img.shape
    xsize2 = (dim[0] - xsize) / 2
    xsize3 = xsize2+1
    ysize2 = (dim[1] - ysize) / 2
    ysize3 = ysize2+1

    # Move
    translation_matrix = np.float32([[1, 0, ysize2], [0, 1, xsize2]]) # translation direction: starts with y
    translated_image = cv2.warpAffine(img, translation_matrix, dim)
    translated_image_2 = translated_image.copy()

    # Background
    array = np.array(img)
    mean_bg = int(np.average(array[1]))
    translated_image_2[np.where(translated_image < 2)] = mean_bg

    # Edge gray; opencv makes by default black corners
    #    left site
    translated_image_2[int(xsize2):int(xsize3), 1:dim[1]] = mean_bg
    #    upper site
    translated_image_2[1:dim[0], int(ysize2):int(ysize3)] = mean_bg
    #    right site
    translated_image_2[int(xsize2+xsize):int(xsize3+xsize), 1:dim[1]] = mean_bg
    #    lower site
    translated_image_2[1:dim[0], int(ysize2+ysize):int(ysize3+ysize)] = mean_bg
    return translated_image_2


def f(x):
    # individual process algorithm, for bigger contrast
    y = np.rint((np.exp(3 * x) - np.sin(x) - 1) * 255 / (np.exp(3) - np.sin(1) - 1))
    return y


def img_proc_functional(img):
    # function resize Img between 0 and 255 and applies function
    img = img / 255
    img = np.array([f(xi) for xi in img])
    return np.asarray(img, dtype="uint8")


def image_processing(path_open, l_filenames, path_save, functional=True):
    """
    Moves to path_open and processes every single image in l_filenames which get saved in path_save.
    The functional processing mode is optional, because of RBC - analysis (which is done in BF Mode).
    There is a conversion to .png data type.
    parameter:
    path_open:              path with images which should be processed
    l_filenames:            list of filenames in path_open, which should be processed
    path_save:              path where images should be saved
    functionl:              optional argument which determines, if functional processing should be done
    """
    # change path
    path_now = os.getcwd()
    os.chdir(path_open)

    for image in l_filenames:
        try:
            # loads tif, png, jpg in grayscale
            img = cv2.imread(image, 0)

            # Translation to the middle
            img = shape(img)

            # Apply Function
            if functional:
                img = img_proc_functional(img)

            # Saving
            os.chdir(path_save)
            image = str(image) + ".png"
            cv2.imwrite(image, img, [cv2.IMWRITE_PNG_COMPRESSION, 0])
            os.chdir(path_open)

        except FileNotFoundError:
            error_input_images()


    os.chdir(path_now)


# load image; uses png image
def tf_load_img(file):
    # standard tensorflow input, with normalisation between 0 and 1
    img_tensor = tf.io.read_file(file)
    img_tensor = tf.image.decode_png(img_tensor, channels=1)
    img_tensor = tf.image.resize(img_tensor, [60,60])
    img_tensor = img_tensor/255
    return img_tensor


# =======================
# prediction analysis and statistics

def prediction_analysis(predictions,  l_images, path_copy, path_save=os.getcwd()):
    """ This function is designed for analysis of predictions. It saves the statistic in txt/ excel files.
    The path is the current work dir, if nothing else is named. Images are saved in separate folders.
    The return of this function depends on analysis parameter.

    parameter:
    predictions :           predictions from TensorFlow
    l_images :              list of image file names; used for predictions; (usually png format)
    path_copy :             path where raw images are stored
    path_save :             directory where statistics are saved; default is folder where program was opened

    return:
    int_ * :                Number of different aggregation types
    l_aggregates:           list of aggregate names
    d_aggregates:           dictionary of aggregate name and aggregation type """
    # =======================
    # paths
    #    get current path
    path_current = os.getcwd()
    #    path save: where statistics and images are saved
    os.chdir(path_save)

    # =======================
    # code for optional analysis
    #    make folders for saving sorted images
    if var_images.get():
        #    mak new images folder
        folder_name = "images_" + str(entry_list[EntryCounter])  # ID which is given in input
        path_save_sorted_images = make_dir(path_save, folder_name)

        #    new sub-folders for each parameter
        os.chdir(path_save_sorted_images)
        parameters = ["single", "duplicate", "triplicate", "multiple"]
        for parameter in parameters:
            os.mkdir(parameter)

        if var_advanced.get():
            os.mkdir("RBC")

    #    list of aggregate- file names; for advanced analysis
    if var_advanced.get():
        l_aggregates = []
        d_aggregates = {}

    # =======================
    # initialization
    #    make counter for number of predictions per group
    int_sing = 0
    int_doub = 0
    int_trip = 0
    int_mult = 0

    #    initialize counter
    z = 0

    while z != len(predictions):
        # format prediction
        a_prediction = np.asarray(predictions[z])
        int_prediction = np.argmax(a_prediction)

        # group length; copy to dir
        if int_prediction == 0:
            int_sing += 1

            if var_images.get():
                # save image
                single_image = l_images[z][:-4]
                sing_path_open = os.path.join(path_copy, single_image)
                sing_path_save = os.path.join(path_save_sorted_images, "single", single_image)
                shutil.copy(sing_path_open, sing_path_save)

        elif int_prediction == 1:
            int_doub += 1
            if var_advanced.get():
                l_aggregates.append(l_images[z][:-4])
                d_aggregates[l_images[z][:-4]] = 1

            if var_images.get():
                # save image
                single_image = l_images[z][:-4]
                doub_path_open = os.path.join(path_copy, single_image)
                doub_path_save = os.path.join(path_save_sorted_images, "duplicate", single_image)
                shutil.copy(doub_path_open, doub_path_save)

        elif int_prediction == 2:
            int_trip += 1
            if var_advanced.get():
                l_aggregates.append(l_images[z][:-4])
                d_aggregates[l_images[z][:-4]] = 2

            if var_images.get():
                # save image
                single_image = l_images[z][:-4]
                trip_path_open = os.path.join(path_copy, single_image)
                trip_path_save = os.path.join(path_save_sorted_images, "triplicate", single_image)
                shutil.copy(trip_path_open, trip_path_save)

        elif int_prediction == 3:
            int_mult += 1
            if var_advanced.get():
                l_aggregates.append(l_images[z][:-4])
                d_aggregates[l_images[z][:-4]] = 3

            if var_images.get():
                # save image
                single_image = l_images[z][:-4]
                mult_path_open = os.path.join(path_copy, single_image)
                mult_path_save = os.path.join(path_save_sorted_images, "multiple", single_image)
                shutil.copy(mult_path_open, mult_path_save)

        z += 1

    # =======================
    # calculate WPA
    nr_all = int_sing + int_doub + int_trip + int_mult

    pro_sing = (int_sing/nr_all)*100
    pro_doub = (int_doub/nr_all)*100
    pro_trip = (int_trip/nr_all)*100
    pro_mult = (int_mult/nr_all)*100

    WPA = (pro_sing + 2*pro_doub + 3*pro_trip + 4*pro_mult)-100

    # =======================
    # output

    if var_advanced.get() and var_images.get():
        # need this information for advanced analysis and image sorting
        os.chdir(path_current)
        return l_aggregates, d_aggregates, int_sing, int_doub, int_trip, int_mult, path_save_sorted_images

    elif var_advanced.get():
        # need this information for advanced analysis
        os.chdir(path_current)
        return l_aggregates, d_aggregates, int_sing, int_doub, int_trip, int_mult

    else:
        # general analysis
        #    statistic is saved next to images folder
        os.chdir(path_save)
        l_statistic_path_files = os.listdir()

        # text - analysis file
        if var_text.get():
            # make analysis file
            if "plateaggrate_analysis.txt" in l_statistic_path_files:
                pass
            else:
                with open("plateaggrate_analysis.txt", "w") as filehandle:
                    print("ID", "time", "WPA", "Nr. Single", "Nr. Doublet", "Nr. Triplet",
                          "Nr. Multiple", sep="\t", file=filehandle)

            # save to analysis file
            with open("plateaggrate_analysis.txt", "a") as filehandle:
                time = str(datetime.datetime.now().strftime("%Y%m%d-%H%M%S"))
                print(entry_list[EntryCounter], time, WPA, int_sing, int_doub,
                      int_trip, int_mult, sep="\t", file=filehandle)

        # excel - analysis file
        if var_excel.get():
            # load/make analysis file and header
            if "plateaggrate_analysis.xlsx" in l_statistic_path_files:
                wb = load_workbook(r"plateaggrate_analysis.xlsx")
                ws = wb.active
            else:
                args = ["ID", "time", "WPA",  "Nr. Single", "Nr. Doublet", "Nr. Triplet", "Nr. Multiple"]
                wb = Workbook()
                ws = wb.active
                for counter in range(len(args)):
                    ws.cell(row=1, column=counter+1).value = args[counter]

            # save to analysis file
            time = str(datetime.datetime.now().strftime("%Y%m%d-%H%M%S"))
            row_number = len_row(ws) + 1
            ws.cell(row=row_number, column=1).value = entry_list[EntryCounter]
            ws.cell(row=row_number, column=2).value = time
            ws.cell(row=row_number, column=3).value = WPA
            ws.cell(row=row_number, column=4).value = int_sing
            ws.cell(row=row_number, column=5).value = int_doub
            ws.cell(row=row_number, column=6).value = int_trip
            ws.cell(row=row_number, column=7).value = int_mult

            wb.save("plateaggrate_analysis.xlsx")
    os.chdir(path_current)


# red blood cell analysis

def prediction_analysis_rbc(predictions, l_aggregate_names, d_aggregates, int_sing, int_doub, int_trip, int_mult,
                            path_copy= os.getcwd(), path_save=os.getcwd()):
    """ This function is designed for analysis of predictions, when used RBC exclusion.
    It saves the statistic in txt/ excel files. These files are saved to the current work dir, if nothing else is named.
    Images are saved in separate folders.

    parameter:
    predictions :           Array List of predictions from TensorFlow: classification between platelets and RBC
    l_aggregate_names :     list of image names; in same order as predictions; aggregates only
    d_aggregates :          Filename (.tif) with prediction (e.g.: {[Name.tif : 1],...} )
    int_ * :                Number of different aggregation type classifications
    path_copy :             sorted images dir with aggregation - subfolders
    path_save :             directory where statistics are saved; default is folder where program was opened"""
    # path current is current path
    path_current = os.getcwd()

    # =======================
    # directory stuff
    #    path save: where statistics and images are saved
    os.chdir(path_save)

    # =======================
    # initialization
    #    count all RBC's
    int_rbc = 0

    z = 0

    while z != len(predictions):
        # format prediction
        a_prediction = np.asarray(predictions[z])
        int_prediction = np.argmax(a_prediction)

        if int_prediction == 1:
            int_rbc += 1

            if d_aggregates[l_aggregate_names[z]] == 1:
                int_doub = int_doub - 1

                if var_images.get():
                    # move image
                    trip_path_open = os.path.join(path_copy,"duplicate", l_aggregate_names[z])
                    trip_path_save = os.path.join(path_copy, "RBC", l_aggregate_names[z])
                    shutil.move(trip_path_open, trip_path_save)

            elif d_aggregates[l_aggregate_names[z]] == 2:
                int_trip = int_trip - 1

                if var_images.get():
                    # move image
                    trip_path_open = os.path.join(path_copy, "triplicate", l_aggregate_names[z])
                    trip_path_save = os.path.join(path_copy, "RBC", l_aggregate_names[z])
                    shutil.move(trip_path_open, trip_path_save)

            elif d_aggregates[l_aggregate_names[z]] == 3:
                int_mult = int_mult - 1

                if var_images.get():
                    # move image
                    trip_path_open = os.path.join(path_copy, "multiple", l_aggregate_names[z])
                    trip_path_save = os.path.join(path_copy, "RBC", l_aggregate_names[z])
                    shutil.move(trip_path_open, trip_path_save)
        z += 1

    # =======================
    # calculate WPA
    nr_all = int_sing + int_doub + int_trip + int_mult

    pro_sing = (int_sing/nr_all)*100
    pro_doub = (int_doub/nr_all)*100
    pro_trip = (int_trip/nr_all)*100
    pro_mult = (int_mult/nr_all)*100

    WPA = (pro_sing + 2*pro_doub + 3*pro_trip + 4*pro_mult)-100

    # =======================
    # output

    # general analysis
    #    statistic is saved next to images_folder
    os.chdir(path_save)
    l_statistic_path_files = os.listdir()

    # text - analysis file
    if var_text.get():
        # make analysis file
        if "plateaggrate_analysis.txt" in l_statistic_path_files:
            pass
        else:
            with open("plateaggrate_analysis.txt", "w") as filehandle:
                print("ID", "time", "WPA", "Nr. Single", "Nr. Doublet", "Nr. Triplet",
                      "Nr. Multiple", "Nr. RBC", sep="\t", file=filehandle)

        # save to analysis file
        with open("plateaggrate_analysis.txt", "a") as filehandle:
            time = str(datetime.datetime.now().strftime("%Y%m%d-%H%M%S"))
            print(entry_list[EntryCounter], time, WPA, int_sing, int_doub,
                  int_trip, int_mult, int_rbc, sep="\t", file=filehandle)

    # excel - analysis file
    if var_excel.get():
        # load/make analysis file and header
        if "plateaggrate_analysis.xlsx" in l_statistic_path_files:
            wb = load_workbook(r"plateaggrate_analysis.xlsx")
            ws = wb.active
        else:
            args = ["ID", "time", "WPA",  "Nr. Single", "Nr. Doublet", "Nr. Triplet", "Nr. Multiple", "Nr. RBC"]
            wb = Workbook()
            ws = wb.active
            for counter in range(len(args)):
                ws.cell(row=1, column=counter+1).value = args[counter]

        # save to analysis file
        time = str(datetime.datetime.now().strftime("%Y%m%d-%H%M%S"))
        row_number = len_row(ws) + 1
        ws.cell(row=row_number, column=1).value = entry_list[EntryCounter]
        ws.cell(row=row_number, column=2).value = time
        ws.cell(row=row_number, column=3).value = WPA
        ws.cell(row=row_number, column=4).value = int_sing
        ws.cell(row=row_number, column=5).value = int_doub
        ws.cell(row=row_number, column=6).value = int_trip
        ws.cell(row=row_number, column=7).value = int_mult
        ws.cell(row=row_number, column=8).value = int_rbc

        wb.save("plateaggrate_analysis.xlsx")
    os.chdir(path_current)

# ==============================================
# execution


def action(path_open, path_save):
    """ This Function structures the whole Program. It's the command of the "run analysis" - button.
    parameter:
    path_open:              List with all paths, where raw images are saved which are used for classification
    path_save:              Path, where Information gets saved (statistics and sorted images)

    paths: (for understanding)
    path:                   single path of one folder which is in path_open
    path_processed:         Path where processed Images get saved (explicit dir)
    """
    # =======================
    # check for input's
    #    no analysis data/ empty strings, if used before (remove it from list)
    if len(path_open) == 0:
        error_paths()
        return
    elif len(path_open) >= 1:
        path_open = [single_path for single_path in path_open if not single_path == ""]

    #    path_save button was one time used, but now empty
    if path_save == "":
        error_paths()
        return

    # path_save button wasn't used before
    try:
        path_save
    except:
        error_load_nn()
        return

    # =======================
    # start
    path_home = os.getcwd()

    # Entry counter for ID's and each run
    global EntryCounter
    EntryCounter = 0

    for path in path_open:
        # ==============================================
        # platelet analysis
        print("Plateaggrate analysis: start")

        # =======================
        # Image Processing
        #   make Dir and get saving path
        path_processed = make_dir(path_save)

        #   raw images
        l_raw_file_names = get_images_list(path)

        #   Processed Images
        print("Plateaggrate analysis: image processing")
        image_processing(path, l_raw_file_names, path_processed)

        # =======================
        # build Datasets (ds)
        os.chdir(path_processed)
        l_images = get_images_list(os.getcwd())

        #   Features: generate Dataset from File Names; load it into the dataset with load function
        ds_feature = tf.data.Dataset.from_tensor_slices(l_images)
        ds_feature = ds_feature.map(tf_load_img)

        #   Labels: generate a list with labels from all file names; convert it to dataset
        ds_label = tf.data.Dataset.from_tensor_slices(l_images)

        #   ds: zip Features, Labels together
        ds = tf.data.Dataset.zip((ds_feature, ds_label))

        #   ds_img_label shape: ((60 [Pixel],60 [Pixel],1 [GRAYSCALE]), ()), types: (tf.float32, tf.int32)
        # =======================
        # final dataset processing
        #   Batch size
        batch_size = 20

        #   final dataset; batching
        ds = ds.batch(batch_size)
        ds = ds.prefetch(buffer_size=40)

        # =======================
        # model
        os.chdir(path_home)
        try:
            model = tf.keras.models.load_model("Aggregate.h5")
        except IOError:
            error_load_nn()

        # =======================
        # make predictions; Array list
        print("Plateaggrate analysis: classification")
        os.chdir(path_processed)
        predictions = model.predict(ds)

        # =======================
        # check predictions
        print("Plateaggrate analysis: statistics")

        if var_advanced.get() and var_images.get():
            l_aggregate_names, d_aggregates, sing, doub, trip, mult, path_sorted_img = prediction_analysis(predictions,
                                                                                                           l_images, path,
                                                                                                           path_save=path_save)
        elif var_advanced.get():
            l_aggregate_names, d_aggregates, sing, doub, trip, mult = prediction_analysis(predictions,
                                                                                          l_images, path, path_save=path_save)
        else:
            prediction_analysis(predictions, l_images, path, path_save=path_save)

        # =======================
        # remove processed images folder
        os.chdir(path_save)
        if os.path.isdir(path_processed):
            shutil.rmtree(path_processed, ignore_errors=True)

        # ==============================================
        # Red Blood Cell analysis
        if var_advanced.get():

            #    move to BF - path
            os.chdir(path)

            # =======================
            # Image Processing
            #   make Dir and get saving path
            path_processed = make_dir(path_save)

            #   Process Images
            #   BF images
            print("Plateaggrate analysis: image processing")
            image_processing(path, l_aggregate_names, path_processed, functional=False)

            os.chdir(path_processed)
            l_images = get_images_list(os.getcwd())

            # =======================
            # build Datasets (ds)
            #   Features: generate Dataset from File Names; load it into the dataset with load function
            ds_Feature = tf.data.Dataset.from_tensor_slices(l_images)
            ds_Feature = ds_Feature.map(tf_load_img)

            #   Labels: generate a list with labels from all file names; convert it to dataset
            ds_Label = tf.data.Dataset.from_tensor_slices(l_images)

            #   ds: zip Features, Labels together
            ds = tf.data.Dataset.zip((ds_Feature, ds_Label))

            #   ds_img_label shape: ((60 [Pixel],60 [Pixel],1 [GRAYSCALE]), ()), types: (tf.float32, tf.int32)
            os.chdir(path_home)

            # =======================
            # final dataset processing
            #   Batch size
            batch_size = 20

            #   final dataset; batching
            ds = ds.batch(batch_size)
            ds = ds.prefetch(buffer_size=40)

            # =======================
            # model
            try:
                model = tf.keras.models.load_model("RBC.h5")
            except IOError:
                error_load_nn()

            # =======================
            # make predictions; Array list
            os.chdir(path_processed)
            print("Plateaggrate analysis: classification")
            predictions = model.predict(ds)

            # =======================
            # check predictions
            print("Plateaggrate analysis: statistics")
            if var_images.get():
                prediction_analysis_rbc(predictions, l_aggregate_names, d_aggregates,
                                        sing, doub, trip, mult, path_sorted_img, path_save=path_save)
            else:
                prediction_analysis_rbc(predictions, l_aggregate_names, d_aggregates,
                                        sing, doub, trip, mult, path_copy=path_processed, path_save=path_save)

        # =======================
        # remove processed images folder
        os.chdir(path_save)
        if os.path.isdir(path_processed):
            shutil.rmtree(path_processed, ignore_errors=True)

        # =======================
        # Entry counter for ID's and each run
        EntryCounter += 1

    os.chdir(path_home)


# ==============================================
# tk functions
#   functions for buttons

def path_open1():
    """ Open dir path """
    global PathOpen1
    PathOpen1 = StringVar()
    input_path = filedialog.askdirectory()
    PathOpen1.set(input_path)

    # show
    text1.delete("1.0", END)
    text1.insert(INSERT, str(PathOpen1.get()))

    # input for entry
    basename1 = os.path.basename(PathOpen1.get())
    entry1.delete(0, END)
    entry1.insert(10, basename1)


def path_open2():
    """ Open dir path """
    global PathOpen2
    PathOpen2 = StringVar()
    input_path = filedialog.askdirectory()
    PathOpen2.set(input_path)

    # show
    text2.delete("1.0", END)
    text2.insert(INSERT, str(PathOpen2.get()))

    # input for entry
    basename2 = os.path.basename(PathOpen2.get())
    entry2.delete(0, END)
    entry2.insert(10, basename2)


def path_open3():
    """ Open dir path """
    global PathOpen3
    PathOpen3 = StringVar()
    input_path = filedialog.askdirectory()
    PathOpen3.set(input_path)

    # show
    text3.delete("1.0", END)
    text3.insert(INSERT, str(PathOpen3.get()))

    # input for entry
    basename3 = os.path.basename(PathOpen3.get())
    entry3.delete(0, END)
    entry3.insert(10, basename3)


def path_open4():
    """ Open dir path """
    global PathOpen4
    PathOpen4 = StringVar()
    input_path = filedialog.askdirectory()
    PathOpen4.set(input_path)

    # show
    text4.delete("1.0", END)
    text4.insert(INSERT, str(PathOpen4.get()))

    # input for entry
    basename4 = os.path.basename(PathOpen4.get())
    entry4.delete(0, END)
    entry4.insert(10, basename4)


def path_open5():
    """ Open dir path """
    global PathOpen5
    PathOpen5 = StringVar()
    input_path = filedialog.askdirectory()
    PathOpen5.set(input_path)

    # show
    text5.delete("1.0", END)
    text5.insert(INSERT, str(PathOpen5.get()))

    # input for entry
    basename5 = os.path.basename(PathOpen5.get())
    entry5.delete(0, END)
    entry5.insert(10, basename5)


def path_open_function():
    """ takes all path_open"""
    path_list = []
    global entry_list
    entry_list = []
    if str(PathOpen1.get()):
        path_list.append(str(PathOpen1.get()))
        entry_list.append(entry1.get())
    if str(PathOpen2.get()):
        path_list.append(str(PathOpen2.get()))
        entry_list.append(entry2.get())
    if str(PathOpen3.get()):
        path_list.append(str(PathOpen3.get()))
        entry_list.append(entry3.get())
    if str(PathOpen4.get()):
        path_list.append(str(PathOpen4.get()))
        entry_list.append(entry4.get())
    if str(PathOpen5.get()):
        path_list.append(str(PathOpen5.get()))
        entry_list.append(entry5.get())

    return path_list


def path_save_function():
    """ This function gets directory path for saving the output."""
    global path_save
    path_save = StringVar()
    input_path = filedialog.askdirectory()
    path_save.set(input_path)

    # show
    text6.delete("1.0", END)
    text6.insert(INSERT, str(path_save.get()))


# messagebox
#    messagebox: opens separate window, with informations for use
def messagebox_action():
    m_text = """Plateaggrate: Measure Platelet Aggreagation
    Step 1: Choose directory with platelet-images (folder).
    Step 2: Choose directory for saving. 
    Step 3: Select Output -format (excel, text, images). 
    Step 4: Run the program (this takes a moment and depends on image number)
    
    info: red blood cell (RBC) exclusion is only done for aggregates."""
    messagebox.showinfo(message=m_text, title="Info/Help")

# errors
def error_paths():
    error_text = """There is no directory path for analysis/ saving.
                    Check:
                    - please insert path"""
    messagebox.showerror(message=error_text, title= "Error")


def error_input_images():
    error_text = """Invalid Input: It's not possible to interpret images. 
                    Check:
                    - directory path (May be empty folder)
                    - images datatype (Can interpret tiff, jpg, png- images)"""
    messagebox.showerror(message=error_text, title="Error")


def error_load_nn():
    error_text = """It's not possible to load NN.
                    Check:
                    - correct .h5 files in plateaggrate.py directory"""
    messagebox.showerror(message=error_text, title= "Error")
# ==============================================
# Graphical User interface


# open main window
main = Tk()

# window name
main.title("Plateaggrate - Platet Aggregation Analysis")

# icon
icon = PhotoImage(file="icon.gif")
main.tk.call('wm', 'iconphoto', main, icon)

# vars
#    paths
PathOpen1 = StringVar()
PathOpen2 = StringVar()
PathOpen3 = StringVar()
PathOpen4 = StringVar()
PathOpen5 = StringVar()

#    outputs
var_excel = BooleanVar(value=True)
var_text = BooleanVar(value=False)
var_images = BooleanVar(value=False)

#    options
var_advanced = BooleanVar(value=False)

# =======================
# tk - widgets

#   buttons
b_1 = Button(main, text="open", command=path_open1, width=14)
b_2 = Button(main, text="open", command=path_open2, width=14)
b_3 = Button(main, text="open", command=path_open3, width=14)
b_4 = Button(main, text="open", command=path_open4, width=14)
b_5 = Button(main, text="open", command=path_open5, width=14)
b_6 = Button(main, text="save", command=path_save_function, width=14)
b_7 = Button(main, text="run analysis", command=lambda: action(path_open_function(), path_save.get()), width=14)
b_8 = Button(main, text="info", command=messagebox_action, width=14)

# =======================
#   Checkbutton

Checkbutton1 = Checkbutton(main, text="excel sheet", variable=var_excel)
Checkbutton2 = Checkbutton(main, text="text file", variable=var_text)
Checkbutton3 = Checkbutton(main, text="images", variable=var_images)
Checkbutton4 = Checkbutton(main, text="RBC exclusion", variable=var_advanced)

# =======================
#   textbox
text1 = Text(main, height=2, width=45)  # open path
text2 = Text(main, height=2, width=45)  # open path
text3 = Text(main, height=2, width=45)  # open path
text4 = Text(main, height=2, width=45)  # open path
text5 = Text(main, height=2, width=45)  # open path
text6 = Text(main, height=2, width=45)  # saving path

S1 = Scrollbar(main)
S2 = Scrollbar(main)
S3 = Scrollbar(main)
S4 = Scrollbar(main)
S5 = Scrollbar(main)
S6 = Scrollbar(main)

# =======================
#   Entrys; used for ID
entry1 = Entry(main, bd=3, width=20, justify=CENTER)
entry2 = Entry(main, bd=3, width=20, justify=CENTER)
entry3 = Entry(main, bd=3, width=20, justify=CENTER)
entry4 = Entry(main, bd=3, width=20, justify=CENTER)
entry5 = Entry(main, bd=3, width=20, justify=CENTER)

# =======================
#   logo/ label
photo = PhotoImage(file="logo.gif")
label_photo = Label(image=photo)

label_ID = Label(main, text="ID", font="bold", width="10")
label_Open = Label(main, text="open", font="bold", width="10")
label_Path = Label(main, text="path", font="bold", width="10")
label_Output = Label(main, text="output", font="bold", width="10")
label_Save = Label(main, text="save", font="bold", width="10")

# ==============================================
# place widgets on window
#   buttons
b_1.grid(row=3, column=2)
b_2.grid(row=4, column=2)
b_3.grid(row=5, column=2)
b_4.grid(row=6, column=2)
b_5.grid(row=7, column=2)
b_6.grid(row=9, column=2)
b_7.grid(row=6, column=6)
b_8.grid(row=9, column=6)

# =======================
#   textbox
text1.grid(row=3, column=3)  # open path
text2.grid(row=4, column=3)  # open path
text3.grid(row=5, column=3)  # open path
text4.grid(row=6, column=3)  # open path
text5.grid(row=7, column=3)  # open path
text6.grid(row=9, column=3)  # saving path

#   scrollbar
S1.grid(row=3, column=4)
S1.config(command=text1.yview)
text1.config(yscrollcommand=S1.set)

S2.grid(row=4, column=4)
S2.config(command=text2.yview)
text2.config(yscrollcommand=S2.set)

S3.grid(row=5, column=4)
S3.config(command=text3.yview)
text3.config(yscrollcommand=S3.set)

S4.grid(row=6, column=4)
S4.config(command=text4.yview)
text4.config(yscrollcommand=S4.set)

S5.grid(row=7, column=4)
S5.config(command=text5.yview)
text5.config(yscrollcommand=S5.set)

S6.grid(row=9, column=4)
S6.config(command=text6.yview)
text6.config(yscrollcommand=S6.set)

# =======================
#   entrys
entry1.grid(row=3, column=1)
entry2.grid(row=4, column=1)
entry3.grid(row=5, column=1)
entry4.grid(row=6, column=1)
entry5.grid(row=7, column=1)

#   start input
entry1.insert(10, "1")
entry2.insert(10, "2")
entry3.insert(10, "3")
entry4.insert(10, "4")
entry5.insert(10, "5")

# =======================
#   checkbutton
Checkbutton1.grid(row=3, column=6)
Checkbutton2.grid(row=4, column=6)
Checkbutton3.grid(row=5, column=6)
Checkbutton4.grid(row=7, column=6)

# =======================
#   labels/ photo
label_photo.grid(row=1, column=3)

label_ID.grid(row=2, column=1)
label_Open.grid(row=2, column=2)
label_Path.grid(row=2, column=3)
label_Output.grid(row=2, column=6)
label_Save.grid(row=9, column=1)

# =======================
#   separator
ttk.Separator(main, orient=VERTICAL).grid(column=5, row=2, rowspan=8, sticky="ns")
ttk.Separator(main, orient=HORIZONTAL).grid(column=1, row=8, columnspan=6, sticky="ew")

main.mainloop()